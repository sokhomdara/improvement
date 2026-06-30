"""
GPPC QAQC Telegram Bot
======================
Project: Grand Phnom Penh City
Usage: Team sends /report command with photos in group chat.
       Bot collects data using BUTTONS, editing ONE message per flow
       (clean, doesn't spam the group chat with new bubbles).
       Saves to Excel (QAQC template format).

Requirements:
    pip install python-telegram-bot openpyxl

Setup:
    1. Create bot via @BotFather → get BOT_TOKEN
    2. Add bot to your Telegram group as Admin
    3. Set BOT_TOKEN below (or via environment variable)
    4. Run: python gppc_qaqc_bot.py
"""

import os
import io
import logging
import threading
from http.server import HTTPServer, BaseHTTPRequestHandler
from datetime import datetime
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    CallbackQueryHandler, ConversationHandler, ContextTypes, filters,
    PicklePersistence
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

# ─────────────────────────────────────────────
# CONFIG — Edit these before running
# ─────────────────────────────────────────────
BOT_TOKEN  = os.environ.get("BOT_TOKEN", "8660783157:AAEBEb_O-z514MdZspps0rp-IH4yg3cz4NM")
EXCEL_FILE = "GPPC_QAQC_Reports.xlsx"
PHOTO_DIR  = "qaqc_photos"  # temp local cache so we can embed thumbnails in Excel

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# CONVERSATION STATES
# ─────────────────────────────────────────────
(
    PICK_ZONE, PICK_BLOCK, TYPE_UNIT, PICK_FLOOR,
    PICK_HTYPE, TYPE_ACTION, PICK_WORKTYPE, PICK_VENDOR,
    PICK_SUPERVISOR, PICK_ENGINEER,
    PICK_RAISED,
    PHOTO_BEFORE, PHOTO_AFTER, TYPE_COMMENT, CONFIRM,
    # Update flow
    UPD_SELECT, UPD_STATUS, UPD_COMMENT, UPD_PHOTO,
) = range(19)

# Preset quick-pick options (most-used values) — bot still allows custom text entry
ZONES        = [f"{i:02d}" for i in range(1, 51)]  # 01 to 50
BLOCKS       = [f"{i:02d}" for i in range(1, 16)]  # 01 to 15
FLOORS       = ["GF", "1F", "2F", "3F", "RF", "EX", "Exterior", "Underground"]
HTYPES       = ["KG", "QN", "QNA", "QNBII", "TWSE", "LASII", "LBIII", "SHC"]
WORKTYPES    = [("🪟 Finishing", "Finishing"), ("🏗️ Structure", "Structure"), ("⚡ MEP", "MEP"), ("🚧 Infrastructure", "Infrastructure")]
VENDORS      = ["ឡាច ពៅ", "វ៉ាន់ សាគីន", "ថន ផល្លា", "Pholla", "NIPPON", "Dulux"]
DEFECTS_FINISHING = [
    "ឥដ្ឋប៉ោង", "ឥដ្ឋមិនពេញរ៉ង", "ជញ្ជាំងបូកប៉ោង",
    "ជ្រុងមិនត្រង់", "សាច់បូកខ្លុក", "ការ៉ូខ្លុក",
    "ការ៉ូបែក", "ម៉ាបបែក", "ថ្នាំមិនស្អាត"
]
DEFECTS_STRUCTURE = [
    "បេតុងសសុះចេញដែក", "បេតុងប៉ោង", "តំទ្បើងដែកខុសប្លង់"
]
DEFECTS_MEP = [
    "ទុយោលិច", "ឆ្លងភ្លើង"
]
DEFECTS_INFRA = [
    "ផ្លូវខូច", "លូទឹកស្ទះ", "បណ្តាញទឹកលិច", "ទ្រនាប់ខូច", "ជញ្ជាំងព័រទ្រ"
]
SUPERVISORS  = ["C9", "C16", "C32", "C45", "C63", "E6", "E20", "E22"]
ENGINEERS    = ['C1', 'C2', 'C3', 'C4', 'C5', 'C6', 'C7', 'C8', 'C9', 'C10', 'C11', 'C12', 'C13', 'C14', 'C15', 'C16', 'C17', 'C18', 'C19', 'C20', 'C21', 'C22', 'C23', 'C24', 'C25', 'C26', 'C27', 'C28', 'C29', 'C30', 'C31', 'C32', 'C33', 'C34', 'C35', 'C36', 'C37', 'C38', 'C39', 'C40', 'C41', 'C42', 'C43', 'C44', 'C45', 'C46', 'C47', 'C48', 'C49', 'C50', 'C51', 'C52', 'C53', 'C54', 'C55', 'C56', 'C57', 'C58', 'C59', 'C60', 'C61', 'C62', 'C63', 'C64', 'C65', 'C66', 'C67', 'C68', 'C69', 'C70', 'C71', 'C72', 'C73', 'C74', 'C75', 'C76', 'C77', 'C78', 'C79', 'C80', 'C81', 'C82', 'C83', 'C84', 'C85', 'C86', 'C87', 'C88', 'C89', 'C90', 'C91', 'C92', 'C93', 'C94', 'C95', 'C96', 'C97', 'C98', 'C99', 'C100', 'C101', 'C102', 'C103', 'C104', 'C105', 'C106', 'C107', 'C108', 'C109', 'C110', 'C111', 'C112', 'C113', 'C114', 'C115', 'C116', 'C117', 'C118', 'C119', 'C120', 'C121', 'C122', 'C123', 'C124', 'C125', 'C126', 'C127', 'C128', 'C129', 'C130', 'C131', 'C132', 'C133', 'C134', 'C135', 'C136', 'C137', 'C138', 'C139', 'C140', 'C141', 'C142', 'C143', 'C144', 'C145', 'C146', 'C147', 'C148', 'C149', 'C150', 'E1', 'E2', 'E3', 'E4', 'E5', 'E6', 'E7', 'E8', 'E9', 'E10', 'E11', 'E12', 'E13', 'E14', 'E15', 'E16', 'E17', 'E18', 'E19', 'E20', 'E21', 'E22', 'E23', 'E24', 'E25', 'E26', 'E27', 'E28', 'E29', 'E30', 'E31', 'E32', 'E33', 'E34', 'E35', 'E36', 'E37', 'E38', 'E39', 'E40', 'E41', 'E42', 'E43', 'E44', 'E45', 'E46', 'E47', 'E48', 'E49', 'E50']
RAISED_BY    = ["Q2", "Q3", "Q6", "Q7", "Q10", "Q11", "Q12", "Q13", "Q16", "Q17", "Q21", "Q26"]
HOUSE_NUMS   = [f"{i:02d}" for i in range(1, 201)]  # 01 to 200

# ─────────────────────────────────────────────
# EXCEL SETUP
# ─────────────────────────────────────────────
HEADERS = [
    "No", "Date", "Picture Before", "Picture After",
    "Zone", "Block", "# House / Unit", "Floor", "House Type",
    "Action Required", "Defect Raised By", "Supervisor In Charge",
    "Site Engineer in Charge", "Type of Work", "Vendor/ SubCon",
    "Responsible by", "Status", "Type of Comments", "Remark"
]
HEADER_FILL  = PatternFill("solid", start_color="1F3864")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
YELLOW_FILL  = PatternFill("solid", start_color="FFF2CC")
GREEN_FILL   = PatternFill("solid", start_color="E2EFDA")
RED_FILL     = PatternFill("solid", start_color="FCE4D6")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def init_excel():
    if os.path.exists(EXCEL_FILE):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QAQC Reports"
    ws["A1"] = "QUALITY MONITORING AND IMPROVEMENT"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="1F3864")
    ws["A2"] = "PROJECT TITLE: GRAND PHNOM PENH CITY"
    ws["A2"].font = Font(name="Arial", bold=True, size=11)
    ws.merge_cells("A1:S1")
    ws.merge_cells("A2:S2")
    ws.row_dimensions[3].height = 20
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    col_widths = [5,12,15,15,8,8,12,10,12,40,15,15,15,12,18,14,12,15,20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    wb.save(EXCEL_FILE)

def get_next_row_and_no():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    row = ws.max_row + 1
    data_rows = max(0, ws.max_row - 3)
    return row, data_rows + 1

def save_to_excel(data: dict):
    init_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    next_row, report_no = get_next_row_and_no()

    status = data.get("status", "Open")
    fill = YELLOW_FILL if status == "Open" else GREEN_FILL if status == "Closed" else RED_FILL

    before_url   = data.get("photo_before_url", "")
    after_url    = data.get("photo_after_url", "")
    before_local = data.get("photo_before_local", "")
    after_local  = data.get("photo_after_local", "")
    row_data = [
        report_no,
        data.get("date", ""),
        "",
        "",
        data.get("zone", ""),
        data.get("block", ""),
        data.get("unit", ""),
        data.get("floor", ""),
        data.get("htype", ""),
        data.get("action", ""),
        data.get("raised", ""),
        data.get("supervisor", ""),
        data.get("engineer", ""),
        data.get("worktype", ""),
        data.get("vendor", ""),
        "Site",
        status,
        data.get("comment_type", "1"),
        data.get("comment", ""),
    ]
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.fill = fill
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Save photos: always write clickable link first, then try thumbnail on top
    for url, local, col_idx, label in [
        (before_url, before_local, 3, "📷 BEFORE"),
        (after_url,  after_local,  4, "✅ AFTER"),
    ]:
        if url:
            cell = ws.cell(row=next_row, column=col_idx, value=label)
            cell.hyperlink = url
            cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Try to embed thumbnail on top if local file exists
            if local and os.path.exists(local):
                try:
                    embed_photo_in_cell(ws, col_idx, next_row, local, url)
                except Exception:
                    pass
    ws.row_dimensions[next_row].height = 18
    wb.save(EXCEL_FILE)
    return report_no

def update_excel_status(report_no: int, new_status: str, remark: str, photo_after_url: str, photo_after_local: str = ""):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=4):
        if row[0].value == report_no:
            row_idx = row[0].row
            row[16].value = new_status
            row[18].value = remark
            if photo_after_url:
                # Always save as clickable link (reliable on cloud hosting)
                cell = ws.cell(row=row_idx, column=4, value="✅ AFTER")
                cell.hyperlink = photo_after_url
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = THIN_BORDER
                # Also try to embed thumbnail if local file exists
                if photo_after_local and os.path.exists(photo_after_local):
                    try:
                        embed_photo_in_cell(ws, 4, row_idx, photo_after_local, photo_after_url)
                    except Exception:
                        pass
            fill = GREEN_FILL if new_status == "Closed" else YELLOW_FILL
            for r_cell in ws[row_idx]:
                r_cell.fill = fill
            break
    wb.save(EXCEL_FILE)

def get_all_reports():
    if not os.path.exists(EXCEL_FILE):
        return []
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0]:
            rows.append({
                "no": row[0], "date": row[1],
                "zone": row[4], "block": row[5], "unit": row[6],
                "action": row[9], "status": row[16]
            })
    return rows

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def status_emoji(s):
    return {"Open": "🟡", "In Progress": "🔵", "Closed": "🟢"}.get(s, "⚪")

def progress_bar(step, total=13):
    filled = "●" * step
    empty = "○" * (total - step)
    return f"{filled}{empty} {step}/{total}"

def grid_buttons(options, prefix, per_row=3, custom=True):
    """Build inline keyboard grid from a list of strings, plus a custom-entry button."""
    rows = []
    row = []
    for i, opt in enumerate(options, 1):
        row.append(InlineKeyboardButton(opt, callback_data=f"{prefix}:{opt}"))
        if i % per_row == 0:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    if custom:
        rows.append([InlineKeyboardButton("✏️ Type custom...", callback_data=f"{prefix}:__custom__")])
    return InlineKeyboardMarkup(rows)

def summary_text(d: dict, report_no=None) -> str:
    no_str = f"#️⃣ Report No: *{report_no}*\n" if report_no else ""
    return (
        f"📋 *QAQC DEFECT REPORT*\n"
        f"{'─'*28}\n"
        f"{no_str}"
        f"📅 {d.get('date','')}\n"
        f"📍 Zone {d.get('zone','')} / Block {d.get('block','')} {d.get('unit','')}\n"
        f"🏠 {d.get('floor','')} — {d.get('htype','')}\n"
        f"⚠️ {d.get('action','')}\n"
        f"🔧 {d.get('worktype','')} | 🏗️ {d.get('vendor','')}\n"
        f"👷 Sup: {d.get('supervisor','')} | Eng: {d.get('engineer','')}\n"
        f"🔎 Raised by: {d.get('raised','')}\n"
        f"📌 {status_emoji(d.get('status','Open'))} {d.get('status','Open')}\n"
        f"💬 {d.get('comment','—')}"
    )

async def get_photo_url(bot, file_id: str) -> str:
    """Return the file_id itself — best for re-sending photos via Telegram API."""
    return file_id  # file_id works directly with send_photo

async def get_photo_full_url(bot, file_id: str) -> str:
    """Return full https URL — for Excel hyperlinks."""
    try:
        f = await bot.get_file(file_id)
        return f.file_path  # returns full https://api.telegram.org/file/bot.../photo.jpg
    except Exception:
        return ""

async def download_photo(bot, file_id: str, label: str) -> str:
    """Download a Telegram photo to local disk and return its path (for Excel embedding)."""
    os.makedirs(PHOTO_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S%f")
    path = os.path.join(PHOTO_DIR, f"{label}_{ts}.jpg")
    f = await bot.get_file(file_id)
    await f.download_to_drive(path)
    return path

def make_thumbnail(src_path: str, max_size=(160, 120)) -> str:
    """Resize a photo down to a small thumbnail for embedding in Excel. Returns thumb path."""
    try:
        img = PILImage.open(src_path)
        img.thumbnail(max_size)
        thumb_path = src_path.replace(".jpg", "_thumb.jpg")
        img.convert("RGB").save(thumb_path, "JPEG", quality=80)
        return thumb_path
    except Exception as e:
        logger.warning(f"Thumbnail failed: {e}")
        return src_path

def embed_photo_in_cell(ws, col_idx: int, row_idx: int, local_path: str, hyperlink_url: str = ""):
    """Embed a photo thumbnail into a specific Excel cell, with optional click-through hyperlink."""
    if not local_path or not os.path.exists(local_path):
        return False
    try:
        thumb = make_thumbnail(local_path, max_size=(160, 120))
        img = XLImage(thumb)
        img.width = 120
        img.height = 100
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        img.anchor = f"{col_letter}{row_idx}"
        ws.add_image(img)
        if hyperlink_url:
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.hyperlink = hyperlink_url
        ws.row_dimensions[row_idx].height = 80   # taller row
        ws.column_dimensions[col_letter].width = 20  # wider column
        return True
    except Exception as e:
        logger.warning(f"Embed photo failed: {e}")
        return False

async def edit_or_send(update: Update, ctx, text, reply_markup=None):
    """Edit the tracked message if possible, else send new and track it."""
    chat_id = update.effective_chat.id
    msg_id = ctx.user_data.get("flow_msg_id")
    if update.callback_query:
        try:
            await update.callback_query.edit_message_text(
                text, parse_mode="Markdown", reply_markup=reply_markup
            )
            ctx.user_data["flow_msg_id"] = update.callback_query.message.message_id
            return
        except Exception:
            pass
    if msg_id:
        try:
            await ctx.bot.edit_message_text(
                chat_id=chat_id, message_id=msg_id, text=text,
                parse_mode="Markdown", reply_markup=reply_markup
            )
            return
        except Exception:
            pass
    sent = await ctx.bot.send_message(
        chat_id=chat_id, text=text, parse_mode="Markdown", reply_markup=reply_markup
    )
    ctx.user_data["flow_msg_id"] = sent.message_id

# ─────────────────────────────────────────────
# /START & /HELP
# ─────────────────────────────────────────────
# Admin user IDs who can use /restart (add your Telegram user ID here)
ADMIN_IDS = [352178789]  # Dara SOKHOM - GPPC QAQC Admin

# Team member full names for alert messages
# Format: "CODE": ("Full Name", "@telegram_username or None")
# Replace None with real @username when you have it
TEAM_INFO = {
    # QA/QC & HSE Team
    "Q3":  ("OUK Mony Oudom",  None),
    "Q6":  ("HOK Yeksrun",     None),
    "Q7":  ("SOKHOM Dara",     "@sokhomdara007"),
    "Q10": ("KHEM Vortana",    None),
    "Q11": ("SOK Chanty",      None),
    "Q12": ("HUO Huot",        None),
    "Q13": ("KONG Sambath",    None),
    "Q16": ("HUOY Seameng",    None),
    "Q17": ("VORN Vanneth",    None),
    "Q21": ("TLY Sophea",      None),
    "Q25": ("CHHEURN Chamnan", None),
    "Q27": ("IM Teklong",      None),
    "Q28": ("CHHEN Sara",      None),
    "Q36": ("SOEM Vicheka",    None),
    # Site Team
    "AD":  ("TOUCH Vanna",     None),
    "C17": ("Mao Maran",       None),
    "C24": ("MEN Vanna",       None),
    "C27": ("Uch Sereykosal",  None),
    "C53": ("Ngeth Ponlue",    None),
    "C61": ("Nut Hok",         None),
    "C82": ("Nhel Vechettra",  None),
    "C97": ("LONG Rathana",    None),
    "M12": ("Phoeung Vanna",   None),
    "E19": ("NOY MAP",         None),
    "E27": ("SOK CHHOM",       None),
    "E29": ("Sor Prostouch",   None),
    "E55": ("DEL RA",          None),
}

def get_mention(code: str) -> str:
    """Return @username if available, else full name, else code."""
    info = TEAM_INFO.get(code)
    if not info:
        return code
    name, username = info
    if username:
        return f"{username} ({name})"
    return f"*{name}* ({code})"

# Keep for backward compat
SITE_MENTIONS = {k: get_mention(k) for k in TEAM_INFO}
QAQC_MENTIONS = SITE_MENTIONS

async def start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 *QAQC Bot* is ready!\n\n"
        "Commands:\n"
        "• /report — Log a new defect\n"
        "• /update — Update defect status\n"
        "• /list — View all open defects\n"
        "• /export — Get Excel report file\n"
        "• /ping — Check if bot is alive\n"
        "• /restart — Restart the bot (admin only)",
        parse_mode="Markdown"
    )

async def ping(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Check if bot is alive and responding."""
    now = datetime.now().strftime("%d %b %Y %H:%M:%S")
    reports = get_all_reports()
    total = len(reports)
    open_c = sum(1 for r in reports if r["status"] == "Open")
    closed_c = sum(1 for r in reports if r["status"] == "Closed")
    await update.message.reply_text(
        f"🟢 *Bot is ALIVE!*\n"
        f"⏰ Server time: {now}\n"
        f"📊 Reports: {total} total | 🟡 {open_c} open | 🟢 {closed_c} closed\n"
        f"✅ All systems running normally!",
        parse_mode="Markdown"
    )

async def restart_bot(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Restart the bot process — admin only."""
    user_id = update.effective_user.id
    if user_id not in ADMIN_IDS:
        await update.message.reply_text("❌ Only admins can restart the bot.")
        return
    await update.message.reply_text(
        "🔄 *Restarting bot...* \nBot will be back in 10 seconds!",
        parse_mode="Markdown"
    )
    # Stop polling cleanly then exit — Render will auto-restart the process
    await ctx.application.stop()
    await ctx.application.shutdown()
    os._exit(0)

async def help_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await start(update, ctx)

# ─────────────────────────────────────────────
# /REPORT — button-driven, single message edited throughout
# ─────────────────────────────────────────────
async def report_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear()
    ctx.user_data["date"] = datetime.now().strftime("%Y-%m-%d")
    ctx.user_data["status"] = "Open"
    try:
        await update.message.delete()
    except Exception:
        pass
    sent = await update.effective_chat.send_message(
        f"📋 *New Defect Report*\n{progress_bar(1)}\n\nSelect *Zone* (01–50):",
        parse_mode="Markdown",
        reply_markup=grid_buttons(ZONES, "zone", per_row=10, custom=False)
    )
    ctx.user_data["flow_msg_id"] = sent.message_id
    return PICK_ZONE

async def pick_zone(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    val = q.data.split(":", 1)[1]
    if val == "__custom__":
        await edit_or_send(update, ctx, f"{progress_bar(1)}\n\n✏️ Type the *Zone* number:")
        return PICK_ZONE
    ctx.user_data["zone"] = val
    await edit_or_send(
        update, ctx,
        f"Zone: *{val}* ✅\n\n{progress_bar(2)}\n\nSelect *Block*:",
        grid_buttons(BLOCKS, "block", per_row=5)
    )
    return PICK_BLOCK

async def text_fallback_zone(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["zone"] = update.message.text.strip()
    try:
        await update.message.delete()
    except Exception:
        pass
    await edit_or_send(
        update, ctx,
        f"Zone: *{ctx.user_data['zone']}* ✅\n\n{progress_bar(2)}\n\nSelect *Block*:",
        grid_buttons(BLOCKS, "block", per_row=5)
    )
    return PICK_BLOCK

async def pick_block(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    val = q.data.split(":", 1)[1]
    if val == "__custom__":
        await edit_or_send(update, ctx, f"{progress_bar(2)}\n\n✏️ Type the *Block* number:")
        return PICK_BLOCK
    ctx.user_data["block"] = val
    await edit_or_send(
        update, ctx,
        f"Block: *{val}* ✅\n\n{progress_bar(3)}\n\nSelect *House / Unit number* (01–200):",
        grid_buttons(HOUSE_NUMS, "unit", per_row=10, custom=True)
    )
    return TYPE_UNIT

async def text_fallback_block(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["block"] = update.message.text.strip()
    try:
        await update.message.delete()
    except Exception:
        pass
    await edit_or_send(
        update, ctx,
        f"Block: *{ctx.user_data['block']}* ✅\n\n{progress_bar(3)}\n\nSelect *House / Unit number* (01–200):",
        grid_buttons(HOUSE_NUMS, "unit", per_row=10, custom=True)
    )
    return TYPE_UNIT

async def pick_unit(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle unit button tap."""
    q = update.callback_query
    await q.answer()
    val = q.data.split(":", 1)[1]
    if val == "__custom__":
        await edit_or_send(update, ctx, f"{progress_bar(3)}\n\n✏️ Type the *House / Unit number*:")
        return TYPE_UNIT
    ctx.user_data["unit"] = val
    await edit_or_send(
        update, ctx,
        f"Unit: *{val}* ✅\n\n{progress_bar(4)}\n\nSelect *Floor*:",
        grid_buttons(FLOORS, "floor", per_row=4, custom=False)
    )
    return PICK_FLOOR

async def type_unit(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    val = update.message.text.strip()
    ctx.user_data["unit"] = "" if val == "-" else val
    try:
        await update.message.delete()
    except Exception:
        pass
    await edit_or_send(
        update, ctx,
        f"Unit: *{ctx.user_data['unit'] or '—'}* ✅\n\n{progress_bar(4)}\n\nSelect *Floor*:",
        grid_buttons(FLOORS, "floor", per_row=4, custom=False)
    )
    return PICK_FLOOR

async def pick_floor(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    val = q.data.split(":", 1)[1]
    ctx.user_data["floor"] = val
    await edit_or_send(
        update, ctx,
        f"Floor: *{val}* ✅\n\n{progress_bar(5)}\n\nSelect *House Type*:",
        grid_buttons(HTYPES, "htype")
    )
    return PICK_HTYPE

async def pick_htype(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    val = q.data.split(":", 1)[1]
    if val == "__custom__":
        await edit_or_send(update, ctx, f"{progress_bar(5)}\n\n✏️ Type the *House Type*:")
        return PICK_HTYPE
    ctx.user_data["htype"] = val
    kb = []
    row = []
    for i, (label, val2) in enumerate(WORKTYPES, 1):
        row.append(InlineKeyboardButton(label, callback_data=f"work:{val2}"))
        if i % 2 == 0:
            kb.append(row)
            row = []
    if row:
        kb.append(row)
    await edit_or_send(
        update, ctx,
        f"House Type: *{val}* ✅\n\n{progress_bar(6)}\n\nSelect *Type of Work*:",
        InlineKeyboardMarkup(kb)
    )
    return PICK_WORKTYPE

async def text_fallback_htype(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["htype"] = update.message.text.strip()
    try:
        await update.message.delete()
    except Exception:
        pass
    kb = []
    row = []
    for i, (label, wval) in enumerate(WORKTYPES, 1):
        row.append(InlineKeyboardButton(label, callback_data=f"work:{wval}"))
        if i % 2 == 0:
            kb.append(row)
            row = []
    if row:
        kb.append(row)
    await edit_or_send(
        update, ctx,
        f"House Type: *{ctx.user_data['htype']}* ✅\n\n{progress_bar(6)}\n\nSelect *Type of Work*:",
        InlineKeyboardMarkup(kb)
    )
    return PICK_WORKTYPE

async def type_action(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    # Custom typed defect description
    ctx.user_data["action"] = update.message.text.strip()
    try:
        await update.message.delete()
    except Exception:
        pass
    await edit_or_send(
        update, ctx,
        f"Defect: *{ctx.user_data['action']}* ✅\n\n{progress_bar(8)}\n\nSelect *Vendor / SubCon*:",
        grid_buttons(VENDORS, "vendor", per_row=2)
    )
    return PICK_VENDOR

async def pick_worktype(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    val = q.data.split(":", 1)[1]
    ctx.user_data["worktype"] = val
    # Show defect presets based on work type
    if val == "Finishing":
        defects = DEFECTS_FINISHING
    elif val == "Structure":
        defects = DEFECTS_STRUCTURE
    elif val == "MEP":
        defects = DEFECTS_MEP
    else:
        defects = DEFECTS_INFRA
    await edit_or_send(
        update, ctx,
        f"Work Type: *{val}* ✅\n\n{progress_bar(6)}\n\n⚠️ Select *Defect / Action Required*:",
        grid_buttons(defects, "defect", per_row=2)
    )
    return TYPE_ACTION

async def pick_defect(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle defect preset button tap."""
    q = update.callback_query
    await q.answer()
    val = q.data.split(":", 1)[1]
    if val == "__custom__":
        await edit_or_send(update, ctx, f"{progress_bar(6)}\n\n✏️ Type the *Defect* in Khmer or English:")
        return TYPE_ACTION
    ctx.user_data["action"] = val
    await edit_or_send(
        update, ctx,
        f"Defect: *{val}* ✅\n\n{progress_bar(8)}\n\nSelect *Vendor / SubCon*:",
        grid_buttons(VENDORS, "vendor", per_row=2)
    )
    return PICK_VENDOR

async def pick_vendor(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    val = q.data.split(":", 1)[1]
    if val == "__custom__":
        await edit_or_send(update, ctx, f"{progress_bar(8)}\n\n✏️ Type the *Vendor / SubCon* name:")
        return PICK_VENDOR
    ctx.user_data["vendor"] = val
    await edit_or_send(
        update, ctx,
        f"Vendor: *{val}* ✅\n\n{progress_bar(9)}\n\nSelect *Supervisor In Charge*:",
        grid_buttons(SUPERVISORS, "sup")
    )
    return PICK_SUPERVISOR

async def text_fallback_vendor(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["vendor"] = update.message.text.strip()
    try:
        await update.message.delete()
    except Exception:
        pass
    await edit_or_send(
        update, ctx,
        f"Vendor: *{ctx.user_data['vendor']}* ✅\n\n{progress_bar(9)}\n\nSelect *Supervisor In Charge*:",
        grid_buttons(SUPERVISORS, "sup")
    )
    return PICK_SUPERVISOR

async def pick_supervisor(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    val = q.data.split(":", 1)[1]
    if val == "__custom__":
        await edit_or_send(update, ctx, f"{progress_bar(9)}\n\n✏️ Type the *Supervisor* code:")
        return PICK_SUPERVISOR
    ctx.user_data["supervisor"] = val
    await edit_or_send(
        update, ctx,
        f"Supervisor: *{val}* ✅\n\n{progress_bar(10)}\n\nSelect *Site Engineer*:",
        grid_buttons(ENGINEERS, "eng", per_row=6)
    )
    return PICK_ENGINEER

async def text_fallback_supervisor(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["supervisor"] = update.message.text.strip()
    try:
        await update.message.delete()
    except Exception:
        pass
    await edit_or_send(
        update, ctx,
        f"Supervisor: *{ctx.user_data['supervisor']}* ✅\n\n{progress_bar(10)}\n\nSelect *Site Engineer*:",
        grid_buttons(ENGINEERS, "eng", per_row=6)
    )
    return PICK_ENGINEER

async def pick_engineer(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    val = q.data.split(":", 1)[1]
    if val == "__custom__":
        await edit_or_send(update, ctx, f"{progress_bar(10)}\n\n✏️ Type the *Site Engineer* code:")
        return PICK_ENGINEER
    ctx.user_data["engineer"] = val
    await edit_or_send(
        update, ctx,
        f"Engineer: *{val}* ✅\n\n{progress_bar(11)}\n\nSelect *Defect Raised By* (QAQC code):",
        grid_buttons(RAISED_BY, "raised", per_row=4)
    )
    return PICK_RAISED

async def text_fallback_engineer(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["engineer"] = update.message.text.strip()
    try:
        await update.message.delete()
    except Exception:
        pass
    await edit_or_send(
        update, ctx,
        f"Engineer: *{ctx.user_data['engineer']}* ✅\n\n{progress_bar(11)}\n\nSelect *Defect Raised By* (QAQC code):",
        grid_buttons(RAISED_BY, "raised", per_row=4)
    )
    return PICK_RAISED

async def pick_raised(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    val = q.data.split(":", 1)[1]
    if val == "__custom__":
        await edit_or_send(update, ctx, f"{progress_bar(11)}\n\n✏️ Type the *QAQC code*:")
        return PICK_RAISED
    ctx.user_data["raised"] = val
    await edit_or_send(
        update, ctx,
        f"Raised by: *{val}* ✅\n\n{progress_bar(12)}\n\n📷 Send the *BEFORE photo* now\n_(or tap Skip)_",
        InlineKeyboardMarkup([[InlineKeyboardButton("⏭️ Skip photo", callback_data="skip:before")]])
    )
    return PHOTO_BEFORE

async def text_fallback_raised(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["raised"] = update.message.text.strip()
    try:
        await update.message.delete()
    except Exception:
        pass
    await edit_or_send(
        update, ctx,
        f"Raised by: *{ctx.user_data['raised']}* ✅\n\n{progress_bar(12)}\n\n📷 Send the *BEFORE photo* now\n_(or tap Skip)_",
        InlineKeyboardMarkup([[InlineKeyboardButton("⏭️ Skip photo", callback_data="skip:before")]])
    )
    return PHOTO_BEFORE

async def photo_before_skip(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    ctx.user_data["photo_before_url"] = ""
    ctx.user_data["photo_before_local"] = ""
    await edit_or_send(
        update, ctx,
        f"{progress_bar(13)}\n\n💬 Add a *comment* for the team:\n_(or type - to skip)_"
    )
    return TYPE_COMMENT

async def photo_before_received(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    file_id = update.message.photo[-1].file_id
    ctx.user_data["photo_before_url"] = await get_photo_full_url(update.get_bot(), file_id)
    ctx.user_data["photo_before_local"] = await download_photo(update.get_bot(), file_id, "before")
    try:
        await update.message.delete()
    except Exception:
        pass
    await edit_or_send(
        update, ctx,
        f"📷 BEFORE photo received ✅\n\n{progress_bar(13)}\n\n💬 Add a *comment* for the team:\n_(or type - to skip)_"
    )
    return TYPE_COMMENT

async def photo_after_skip(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    ctx.user_data["photo_after_url"] = ""
    await edit_or_send(
        update, ctx,
        f"{progress_bar(13)}\n\n💬 Add a *comment* for the team:\n_(or type - to skip)_"
    )
    return TYPE_COMMENT

async def photo_after_received(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    file_id = update.message.photo[-1].file_id
    url = await get_photo_url(update.get_bot(), file_id)
    ctx.user_data["photo_after_url"] = url
    ctx.user_data["photo_after_local"] = await download_photo(update.get_bot(), file_id, "after")
    try:
        await update.message.delete()
    except Exception:
        pass
    await edit_or_send(
        update, ctx,
        f"✅ AFTER photo received ✅\n\n{progress_bar(13)}\n\n💬 Add a *comment* for the team:\n_(or type - to skip)_"
    )
    return TYPE_COMMENT

async def type_comment(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    val = update.message.text.strip()
    ctx.user_data["comment"] = "" if val == "-" else val
    ctx.user_data["comment_type"] = "1"
    try:
        await update.message.delete()
    except Exception:
        pass
    kb = [[
        InlineKeyboardButton("✅ Confirm & Save", callback_data="confirm:yes"),
        InlineKeyboardButton("❌ Cancel", callback_data="confirm:no"),
    ]]
    await edit_or_send(
        update, ctx,
        summary_text(ctx.user_data) + "\n\nSave this report?",
        InlineKeyboardMarkup(kb)
    )
    return CONFIRM

async def confirm_report(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    if q.data == "confirm:yes":
        init_excel()
        report_no = save_to_excel(ctx.user_data)
        chat_id   = update.effective_chat.id

        # 1. Edit flow message to show summary
        await q.edit_message_text(
            f"✅ *Report #{report_no} saved!*\n\n"
            + summary_text(ctx.user_data, report_no)
            + "\n\n📊 Use /export to download the Excel file.",
            parse_mode="Markdown"
        )

        # 2. Send BEFORE photo in group if available
        before_file_id = ctx.user_data.get("photo_before_local", "")
        before_url     = ctx.user_data.get("photo_before_url", "")
        if before_file_id and os.path.exists(before_file_id):
            try:
                await ctx.bot.send_photo(
                    chat_id=chat_id,
                    photo=open(before_file_id, "rb"),
                    caption=f"📷 *BEFORE photo — Report #{report_no}*",
                    parse_mode="Markdown"
                )
            except Exception:
                pass
        elif before_url:
            try:
                await ctx.bot.send_photo(
                    chat_id=chat_id,
                    photo=before_url,
                    caption=f"📷 *BEFORE photo — Report #{report_no}*",
                    parse_mode="Markdown"
                )
            except Exception:
                pass

        # 3. Build alert with full name + @mention for site team
        sup_code  = ctx.user_data.get("supervisor", "")
        eng_code  = ctx.user_data.get("engineer", "")
        raised    = ctx.user_data.get("raised", "")
        zone      = ctx.user_data.get("zone", "")
        block     = ctx.user_data.get("block", "")
        unit      = ctx.user_data.get("unit", "")
        action    = ctx.user_data.get("action", "")
        worktype  = ctx.user_data.get("worktype", "")
        floor     = ctx.user_data.get("floor", "")
        htype     = ctx.user_data.get("htype", "")

        sup_mention   = get_mention(sup_code)
        eng_mention   = get_mention(eng_code)
        raised_mention = get_mention(raised)

        alert_text = (
            f"🚨 *DEFECT ALERT — Report #{report_no}*\n"
            f"{'─'*28}\n"
            f"📅 {ctx.user_data.get('date','')}\n"
            f"📍 Zone {zone} / Block {block} / Unit {unit}\n"
            f"🏠 {floor} — {htype}\n"
            f"⚠️ *{action}* ({worktype})\n"
            f"{'─'*28}\n"
            f"👷 Supervisor: {sup_mention}\n"
            f"🔧 Engineer: {eng_mention}\n"
            f"🔎 Raised by: {raised_mention}\n"
            f"{'─'*28}\n"
            f"📌 Status: 🟡 Open\n"
            f"⚡ *Please fix and use /update to close!*"
        )
        await ctx.bot.send_message(
            chat_id=chat_id,
            text=alert_text,
            parse_mode="Markdown"
        )

    else:
        await q.edit_message_text("❌ Report cancelled.")
    ctx.user_data.clear()
    return ConversationHandler.END

# ─────────────────────────────────────────────
# /UPDATE — button-driven
# ─────────────────────────────────────────────
async def update_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear()
    reports = get_all_reports()
    open_reports = [r for r in reports if r["status"] != "Closed"]
    if not open_reports:
        await update.message.reply_text("✅ No open defects found.")
        return ConversationHandler.END

    kb_rows, row = [], []
    for i, r in enumerate(open_reports[-15:], 1):
        row.append(InlineKeyboardButton(f"#{r['no']}", callback_data=f"updsel:{r['no']}"))
        if i % 5 == 0:
            kb_rows.append(row); row = []
    if row:
        kb_rows.append(row)

    lines = "\n".join([
        f"{status_emoji(r['status'])} *#{r['no']}* — Z{r['zone']}/B{r['block']} — {str(r['action'])[:35]}..."
        for r in open_reports[-15:]
    ])
    sent = await update.message.reply_text(
        f"📋 *Open Defects:*\n{lines}\n\nTap a report number to update:",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(kb_rows)
    )
    ctx.user_data["flow_msg_id"] = sent.message_id
    try:
        await update.message.delete()
    except Exception:
        pass
    return UPD_SELECT

async def upd_select(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    ctx.user_data["upd_no"] = int(q.data.split(":", 1)[1])
    kb = [[
        InlineKeyboardButton("🟡 Open", callback_data="updstat:Open"),
        InlineKeyboardButton("🔵 In Progress", callback_data="updstat:In Progress"),
        InlineKeyboardButton("🟢 Closed", callback_data="updstat:Closed"),
    ]]
    await edit_or_send(
        update, ctx,
        f"Report #{ctx.user_data['upd_no']} — Select new *Status*:",
        InlineKeyboardMarkup(kb)
    )
    return UPD_STATUS

async def upd_status_chosen(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    ctx.user_data["upd_status"] = q.data.split(":", 1)[1]
    await edit_or_send(
        update, ctx,
        f"Status: *{ctx.user_data['upd_status']}* ✅\n\n💬 Describe what was done to fix the defect:\n_(or type - to skip)_"
    )
    return UPD_COMMENT

async def upd_comment(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    val = update.message.text.strip()
    ctx.user_data["upd_comment"] = "" if val == "-" else val
    try:
        await update.message.delete()
    except Exception:
        pass
    await edit_or_send(
        update, ctx,
        "📷 Send the *AFTER photo* (completion proof)\n_(or tap Skip)_",
        InlineKeyboardMarkup([[InlineKeyboardButton("⏭️ Skip photo", callback_data="updskip:1")]])
    )
    return UPD_PHOTO

async def upd_photo_skip(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    await finish_update(update, ctx, "", "")
    return ConversationHandler.END

async def upd_photo_received(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    file_id = update.message.photo[-1].file_id
    full_url = await get_photo_full_url(update.get_bot(), file_id)
    local_path = await download_photo(update.get_bot(), file_id, f"after_upd{ctx.user_data.get('upd_no','x')}")
    ctx.user_data["upd_photo_file_id"] = file_id  # for sending photo in chat
    try:
        await update.message.delete()
    except Exception:
        pass
    await finish_update(update, ctx, full_url, local_path)
    return ConversationHandler.END

async def finish_update(update: Update, ctx: ContextTypes.DEFAULT_TYPE, photo_url: str, photo_local: str = ""):
    update_excel_status(
        ctx.user_data["upd_no"],
        ctx.user_data["upd_status"],
        ctx.user_data["upd_comment"],
        photo_url,
        photo_local
    )
    report_no   = ctx.user_data["upd_no"]
    new_status  = ctx.user_data["upd_status"]
    comment     = ctx.user_data["upd_comment"] or "—"
    chat_id     = update.effective_chat.id

    text = (
        f"✅ *Report #{report_no} updated!*\n"
        f"Status → {status_emoji(new_status)} *{new_status}*\n"
        f"💬 Comment: {comment}"
    )

    # Edit the flow message to show summary
    await edit_or_send(update, ctx, text)

    # Send AFTER photo separately so it's visible in the group
    file_id = ctx.user_data.get("upd_photo_file_id", "")
    if file_id:
        try:
            await ctx.bot.send_photo(
                chat_id=chat_id,
                photo=file_id,
                caption=f"✅ *AFTER photo — Report #{report_no}*\nStatus: {status_emoji(new_status)} {new_status}",
                parse_mode="Markdown"
            )
        except Exception:
            pass

    ctx.user_data.clear()

# ─────────────────────────────────────────────
# /LIST & /EXPORT
# ─────────────────────────────────────────────
async def list_reports(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    reports = get_all_reports()
    if not reports:
        await update.message.reply_text("📭 No reports yet. Use /report to log the first defect.")
        return
    lines = []
    for r in reports[-20:]:
        lines.append(
            f"{status_emoji(r['status'])} *#{r['no']}* {r['date']} | "
            f"Z{r['zone']}/B{r['block']} {r['unit']} | {str(r['action'])[:35]}..."
        )
    open_c  = sum(1 for r in reports if r["status"] == "Open")
    prog_c  = sum(1 for r in reports if r["status"] == "In Progress")
    close_c = sum(1 for r in reports if r["status"] == "Closed")
    await update.message.reply_text(
        f"📊 *QAQC Report Summary*\n"
        f"Total: {len(reports)} | 🟡 Open: {open_c} | 🔵 Progress: {prog_c} | 🟢 Closed: {close_c}\n"
        f"{'─'*28}\n" + "\n".join(lines),
        parse_mode="Markdown"
    )

async def export_excel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not os.path.exists(EXCEL_FILE):
        await update.message.reply_text("📭 No reports yet.")
        return
    await update.message.reply_document(
        document=open(EXCEL_FILE, "rb"),
        filename=f"GPPC_QAQC_{datetime.now().strftime('%Y%m%d')}.xlsx",
        caption=f"📊 Quality Monitoring Report\n{datetime.now().strftime('%d %b %Y %H:%M')}"
    )

# ─────────────────────────────────────────────
# CANCEL
# ─────────────────────────────────────────────
async def cancel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear()
    await update.message.reply_text("❌ Cancelled. Use /report to start again.")
    return ConversationHandler.END

async def orphaned_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Catches button taps when the bot lost conversation state (e.g. after restart)."""
    q = update.callback_query
    await q.answer("⚠️ Session expired — please start again", show_alert=True)
    try:
        await q.edit_message_text(
            "⚠️ *Session expired* (bot restarted)\n\nPlease type /report to start a new defect, or /update to continue updating a report.",
            parse_mode="Markdown"
        )
    except Exception:
        pass

async def orphaned_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Catches stray text messages when no conversation is active."""
    text = update.message.text.strip().lower()
    # Ignore short replies that look like leftover form answers
    if len(text) <= 15:
        await update.message.reply_text(
            "ℹ️ No active form right now. Type /report to log a defect or /update to update one.",
        )

# ─────────────────────────────────────────────
# KEEP-ALIVE SERVER (prevents Render from sleeping)
# ─────────────────────────────────────────────
class KeepAliveHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.send_header("Content-Type", "text/plain")
        self.end_headers()
        self.wfile.write(b"QAQC Bot is alive! Status: OK")
    def do_HEAD(self):
        self.send_response(200)
        self.send_header("Content-Type", "text/plain")
        self.end_headers()
    def log_message(self, format, *args):
        pass

def keep_alive():
    port = int(os.environ.get("PORT", 8080))
    server = HTTPServer(("0.0.0.0", port), KeepAliveHandler)
    thread = threading.Thread(target=server.serve_forever)
    thread.daemon = True
    thread.start()
    print(f"🌐 Keep-alive server running on port {port}")

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    keep_alive()  # Start web server FIRST so Render health check passes immediately
    init_excel()

    persistence = PicklePersistence(filepath="bot_state.pickle")
    app = (
        ApplicationBuilder()
        .token(BOT_TOKEN)
        .connect_timeout(30)
        .read_timeout(30)
        .write_timeout(30)
        .pool_timeout(30)
        .persistence(persistence)
        .build()
    )

    report_conv = ConversationHandler(
        entry_points=[CommandHandler("report", report_start)],
        name="report_conv",
        persistent=True,
        states={
            PICK_ZONE:       [CallbackQueryHandler(pick_zone, pattern="^zone:"),
                               MessageHandler(filters.TEXT & ~filters.COMMAND, text_fallback_zone)],
            PICK_BLOCK:      [CallbackQueryHandler(pick_block, pattern="^block:"),
                               MessageHandler(filters.TEXT & ~filters.COMMAND, text_fallback_block)],
            TYPE_UNIT:       [CallbackQueryHandler(pick_unit, pattern="^unit:"),
                               MessageHandler(filters.TEXT & ~filters.COMMAND, type_unit)],
            PICK_FLOOR:      [CallbackQueryHandler(pick_floor, pattern="^floor:")],
            PICK_HTYPE:      [CallbackQueryHandler(pick_htype, pattern="^htype:"),
                               MessageHandler(filters.TEXT & ~filters.COMMAND, text_fallback_htype)],
            TYPE_ACTION:     [CallbackQueryHandler(pick_defect, pattern="^defect:"),
                               MessageHandler(filters.TEXT & ~filters.COMMAND, type_action)],
            PICK_WORKTYPE:   [CallbackQueryHandler(pick_worktype, pattern="^work:")],
            PICK_VENDOR:     [CallbackQueryHandler(pick_vendor, pattern="^vendor:"),
                               MessageHandler(filters.TEXT & ~filters.COMMAND, text_fallback_vendor)],
            PICK_SUPERVISOR: [CallbackQueryHandler(pick_supervisor, pattern="^sup:"),
                               MessageHandler(filters.TEXT & ~filters.COMMAND, text_fallback_supervisor)],
            PICK_ENGINEER:   [CallbackQueryHandler(pick_engineer, pattern="^eng:"),
                               MessageHandler(filters.TEXT & ~filters.COMMAND, text_fallback_engineer)],
            PICK_RAISED:     [CallbackQueryHandler(pick_raised, pattern="^raised:"),
                               MessageHandler(filters.TEXT & ~filters.COMMAND, text_fallback_raised)],
            PHOTO_BEFORE:    [CallbackQueryHandler(photo_before_skip, pattern="^skip:before"),
                               MessageHandler(filters.PHOTO, photo_before_received)],
            TYPE_COMMENT:    [MessageHandler(filters.TEXT & ~filters.COMMAND, type_comment)],
            CONFIRM:         [CallbackQueryHandler(confirm_report, pattern="^confirm:")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    update_conv = ConversationHandler(
        entry_points=[CommandHandler("update", update_start)],
        name="update_conv",
        persistent=True,
        states={
            UPD_SELECT:  [CallbackQueryHandler(upd_select, pattern="^updsel:")],
            UPD_STATUS:  [CallbackQueryHandler(upd_status_chosen, pattern="^updstat:")],
            UPD_COMMENT: [MessageHandler(filters.TEXT & ~filters.COMMAND, upd_comment)],
            UPD_PHOTO:   [CallbackQueryHandler(upd_photo_skip, pattern="^updskip:"),
                          MessageHandler(filters.PHOTO, upd_photo_received)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_cmd))
    app.add_handler(CommandHandler("list", list_reports))
    app.add_handler(CommandHandler("export", export_excel))
    app.add_handler(CommandHandler("ping", ping))
    app.add_handler(CommandHandler("restart", restart_bot))
    app.add_handler(report_conv)
    app.add_handler(update_conv)

    # Catch-all fallback handlers (must be added LAST, lowest priority)
    app.add_handler(CallbackQueryHandler(orphaned_callback))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, orphaned_text))

    print("🤖 QAQC Bot is running...")
    print(f"📁 Excel file: {EXCEL_FILE}")

    app.run_polling(poll_interval=0.1, timeout=30, drop_pending_updates=True)

if __name__ == "__main__":
    main()
